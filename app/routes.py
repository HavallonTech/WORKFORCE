from flask import Blueprint, app, current_app, flash, jsonify, render_template, request, redirect, url_for
from flask_login import current_user, login_user, login_required, logout_user
from app.models import (
    Attendance,
    AttendanceSchedule,
    AttendanceScheduleConfig,
    AuditLog, Department, User, Unit, FieldAttendance, UnitLocation, AttendanceScheduleConfig,
    DynamicAttendance, AttendanceBreak)
from app import db
from app.helpers import toast, generate_staff_id
from sqlalchemy import or_
from app.utils.audit import log_audit
import re
from datetime import (
    datetime,
    timedelta,
    date
)
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from werkzeug.utils import secure_filename


from openpyxl import Workbook, load_workbook
import math
from app.utils.timezone import nigeria_now
import os
import base64
from app.utils.permissions import (
    admin_required,
    superadmin_required
)

from app.utils.geofence import (
    distance_in_meters
)
from io import BytesIO
import pandas as pd
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors

from flask import send_file

from app.utils.field_attendance import (
    get_current_punch
)
from app.models import (
    FieldAttendance
)
from app.utils.attendance import (
    get_current_checkpoint
)

main = Blueprint("main", __name__)



@main.before_app_request
def create_super_admin():

    if User.query.filter_by(role="superadmin").first():
        return

    admin = User(
        username="superadmin",
        email="kestplanet@gmail.com",
        role="superadmin"
    )

    admin.set_password("Admin@123")

    db.session.add(admin)

    db.session.commit()

    print("Super Admin Created")

@main.route("/", methods=["GET", "POST"])
def login():

    if current_user.is_authenticated:
        return redirect(url_for("main.dashboard"))

    if request.method == "POST":

        username = request.form.get("username")
        password = request.form.get("password")

        user = User.query.filter_by(
            username=username
        ).first()

        if not user:
            toast(
                "Invalid username or password",
                "danger"
            )
            return render_template(
                "auth/login.html"
            )

        # Verify password first
        if user.check_password(password):

            if user.must_change_password:

                login_user(user)

                return redirect(
                    url_for(
                        "main.change_password"
                    )
                )

            user.last_login = nigeria_now()

            db.session.commit()

            login_user(user)

            log_audit(
                "Authentication",
                "Login",
                f"{user.username} logged in"
            )

            return redirect(
                url_for("main.dashboard")
            )

        toast(
            "Invalid username or password",
            "danger"
        )

    return render_template(
        "auth/login.html"
    )


@main.route("/dashboard")
@login_required
def dashboard():

    if current_user.role in [

        "admin",
        "superadmin"

    ]:

        return redirect(
            url_for(
                "main.attendance_dashboard_v2"
            )
        )

    today = nigeria_now().date()

    total_checkpoints = 0

    completed_count = DynamicAttendance.query.filter_by(

        user_id=current_user.id,

        attendance_date=today

    ).count()

    break_used = AttendanceBreak.query.filter_by(

        user_id=current_user.id,

        attendance_date=today

    ).first()

    config = AttendanceScheduleConfig.query.filter_by(

        unit_id=current_user.unit_id

    ).first()

    if config:

        total_checkpoints = config.total_checkpoints

    expected = total_checkpoints

    if break_used:

        expected = max(
            0,
            expected - 1
        )

    compliance = round(

        (
            completed_count /
            max(1, expected)
        ) * 100,

        1

    )

    remaining = max(

        0,

        expected - completed_count

    )

    schedule = get_current_checkpoint(

        current_user.unit_id

    )

    recent_attendance = DynamicAttendance.query.filter_by(

        user_id=current_user.id

    ).order_by(

        DynamicAttendance.id.desc()

    ).limit(5).all()

    today_attendance = Attendance.query.filter_by(

        user_id=current_user.id,

        attendance_date=today

    ).first()
    has_active_checkpoint = (

        schedule is not None

    )
    current_time = nigeria_now().strftime(
        "%I:%M %p" )

    return render_template(

        "dashboard/index.html",

        compliance=compliance,
        completed_count=completed_count,
        current_time=current_time,
        has_active_checkpoint=has_active_checkpoint,
        config=config,
        schedule=schedule,
        expected=expected,
        remaining=remaining,
        break_used=break_used,
        today_attendance=today_attendance,
        recent_attendance=recent_attendance

    )

@main.route("/units")
@login_required
@admin_required
def units():

    search = request.args.get(

        "search",

        ""

    ).strip()

    page = request.args.get(

        "page",

        1,

        type=int

    )

    query = Unit.query

    if search:

        query = query.filter(

            db.or_(

                Unit.name.ilike(

                    f"%{search}%"

                ),

                Unit.code.ilike(

                    f"%{search}%"

                ),

                Unit.phone.ilike(

                    f"%{search}%"

                ),

                Unit.email.ilike(

                    f"%{search}%"

                )

            )

        )

    pagination = query.order_by(

        Unit.name

    ).paginate(

        page=page,

        per_page=15,

        error_out=False

    )

    total_projects = Unit.query.count()

    active_projects = Unit.query.filter_by(

        status=True

    ).count()

    inactive_projects = Unit.query.filter_by(

        status=False

    ).count()
    return render_template(

        "units/list.html",

        pagination=pagination,

        projects=pagination.items,

        search=search,

        total_projects=total_projects,

        active_projects=active_projects,

        inactive_projects=inactive_projects

    )


@main.route("/units/add", methods=["GET", "POST"])
@login_required
@admin_required
def add_unit():


    if request.method == "POST":

        code = request.form.get("code").strip().upper()

        existing = Unit.query.filter_by(
            code=code
        ).first()

        if existing:

            toast(
                "Unit code already exists",
                "warning"
            )

            return redirect(
                url_for("main.add_unit")
            )

        unit = Unit(
            code=code,
            name=request.form.get("name"),
            address=request.form.get("address"),
            phone=request.form.get("phone"),
            email=request.form.get("email")
        )

        db.session.add(unit)

        db.session.commit()
        log_audit(
            "Units",
            "Create Unit",
            f"Created Unit {unit.code}"
        )
        toast(
            "Unit created successfully",
            "success"
        )

        return redirect(url_for("main.units"))

    return render_template("units/add.html")

@main.route("/units/edit/<int:id>", methods=["GET", "POST"])
@login_required
@admin_required
def edit_unit(id):

    unit = Unit.query.get_or_404(id)

    if request.method == "POST":

        unit.code =  request.form.get("code").strip().upper()

        unit.name = request.form.get("name")
        unit.address = request.form.get("address")
        unit.phone = request.form.get("phone")
        unit.email = request.form.get("email")

        db.session.commit()

        return redirect(url_for("main.units"))

    return render_template(
        "units/edit.html",
        unit=unit
    )

@main.route("/units/toggle/<int:id>")
@login_required
def toggle_unit(id):

    unit = Unit.query.get_or_404(id)

    unit.status = not unit.status

    db.session.commit()

    return redirect(url_for("main.units"))

@main.route("/test-toast")
def test_toast():

    flash(
        "WorkForce Toast System Working!",
        "success"
    )

    return redirect(
        url_for("main.dashboard")
    )


@main.route("/unit-locations")
@login_required
@admin_required
def unit_locations():

    search = request.args.get(

        "search",

        ""

    ).strip()

    page = request.args.get(

        "page",

        1,

        type=int

    )

    query = UnitLocation.query.join(

        Unit

    )

    if search:

        query = query.filter(

            or_(

                Unit.name.ilike(

                    f"%{search}%"

                ),

                UnitLocation.name.ilike(

                    f"%{search}%"

                ),

                UnitLocation.latitude.cast(

                    db.String

                ).ilike(

                    f"%{search}%"

                ),

                UnitLocation.longitude.cast(

                    db.String

                ).ilike(

                    f"%{search}%"

                )

            )

        )

    pagination = query.order_by(

        UnitLocation.id.desc()

    ).paginate(

        page=page,

        per_page=15,

        error_out=False

    )

    total_locations = UnitLocation.query.count()

    active_locations = UnitLocation.query.filter_by(

        status=True

    ).count()

    inactive_locations = UnitLocation.query.filter_by(

        status=False

    ).count()

    return render_template(

        "unit_locations/list.html",

        locations=pagination.items,

        pagination=pagination,

        search=search,

        total_locations=total_locations,

        active_locations=active_locations,

        inactive_locations=inactive_locations

    )

@main.route(
    "/unit-locations/add",
    methods=["GET", "POST"]
)
@login_required
def add_unit_location():

    units = Unit.query.filter_by(
        status=True
    ).all()

    if request.method == "POST":

        location = UnitLocation(

            unit_id=request.form.get(
                "unit_id"
            ),

            name=request.form.get(
                "name"
            ),

            latitude=request.form.get(
                "latitude"
            ),

            longitude=request.form.get(
                "longitude"
            ),

            radius=request.form.get(
                "radius"
            ),

            status=True
        )

        db.session.add(location)

        db.session.commit()

        log_audit(
            "Unit Locations",
            "Create Location",
            f"Created {location.name}"
        )

        toast(
            "Location added successfully",
            "success"
        )

        return redirect(
            url_for("main.unit_locations")
        )

    return render_template(
        "unit_locations/add.html",
        units=units
    )
@main.route("/unit-locations/toggle/<int:id>")
@login_required
def toggle_unit_location(id):

    location = UnitLocation.query.get_or_404(id)

    location.status = not location.status

    db.session.commit()

    log_audit(
        "Unit Locations",
        "Toggle Location",
        f"{location.name} status changed"
    )

    toast(
        "Location status updated",
        "info"
    )

    return redirect(
        url_for("main.unit_locations")
    )

@main.route(
    "/unit-locations/edit/<int:id>",
    methods=["GET", "POST"]
)
@login_required
def edit_unit_location(id):

    location = UnitLocation.query.get_or_404(id)

    units = Unit.query.filter_by(
        status=True
    ).all()

    if request.method == "POST":

        location.unit_id = request.form.get(
            "unit_id"
        )

        location.name = request.form.get(
            "name"
        )

        location.latitude = request.form.get(
            "latitude"
        )

        location.longitude = request.form.get(
            "longitude"
        )

        location.radius = request.form.get(
            "radius"
        )

        db.session.commit()

        log_audit(
            "Unit Locations",
            "Edit Location",
            f"Updated {location.name}"
        )

        toast(
            "Location updated successfully",
            "success"
        )

        return redirect(
            url_for("main.unit_locations")
        )

    return render_template(
        "unit_locations/edit.html",
        location=location,
        units=units
    )

@main.route("/departments")
@login_required
def departments():

    departments = Department.query.order_by(
        Department.name
    ).all()

    return render_template(
        "departments/list.html",
        departments=departments
    )

@main.route(
    "/departments/add",
    methods=["GET", "POST"]
)
@login_required
def add_department():

    units = Unit.query.filter_by(
        status=True
    ).all()

    if request.method == "POST":

        department = Department(

            unit_id=request.form.get(
                "unit_id"
            ),

            name=request.form.get(
                "name"
            ),

            description=request.form.get(
                "description"
            ),

            status=True
        )

        db.session.add(department)

        db.session.commit()

        log_audit(
            "Departments",
            "Create Department",
            f"Created {department.name}"
        )

        toast(
            "Department created successfully",
            "success"
        )

        return redirect(
            url_for("main.departments")
        )

    return render_template(
        "departments/add.html",
        units=units
    )

@main.route(
    "/departments/edit/<int:id>",
    methods=["GET", "POST"]
)
@login_required
def edit_department(id):

    department = Department.query.get_or_404(id)

    units = Unit.query.filter_by(
        status=True
    ).all()

    if request.method == "POST":

        department.unit_id = request.form.get(
            "unit_id"
        )

        department.name = request.form.get(
            "name"
        )

        department.description = request.form.get(
            "description"
        )

        db.session.commit()

        log_audit(
            "Departments",
            "Edit Department",
            f"Updated {department.name}"
        )

        toast(
            "Department updated successfully",
            "success"
        )

        return redirect(
            url_for("main.departments")
        )

    return render_template(
        "departments/edit.html",
        department=department,
        units=units
    )

@main.route("/departments/toggle/<int:id>")
@login_required
def toggle_department(id):

    department = Department.query.get_or_404(id)

    department.status = not department.status

    db.session.commit()

    log_audit(
        "Departments",
        "Toggle Department",
        f"{department.name} status changed"
    )

    toast(
        "Department status updated",
        "info"
    )

    return redirect(
        url_for("main.departments")
    )

@main.route("/users")
@login_required
def users():

    users = User.query.order_by(
        User.full_name
    ).all()

    total_users = User.query.count()

    active_users = User.query.filter_by(
        is_active_user=True
    ).count()

    inactive_users = User.query.filter_by(
        is_active_user=False
    ).count()

    return render_template(
        "users/list.html",
        users=users,
        total_users=total_users,
        active_users=active_users,
        inactive_users=inactive_users
    )

@main.route(
    "/users/add",
    methods=["GET","POST"]
)
@login_required
@admin_required
def add_user():

    units = Unit.query.filter_by(
        status=True
    ).all()

    departments = Department.query.filter_by(
        status=True
    ).all()

    if request.method == "POST":

        staff_id = request.form.get(
            "staff_id",
            ""
        ).strip()

        username = request.form.get(
            "username",
            ""
        ).strip()

        email = request.form.get(
            "email",
            ""
        ).strip().lower()

        full_name = request.form.get(
            "full_name",
            ""
        ).strip()

        phone = request.form.get(
            "phone",
            ""
        ).strip()

        password = request.form.get(
            "password"
        )

        confirm_password = request.form.get(
            "confirm_password"
        )

        unit_id = request.form.get(
            "unit_id"
        )

        department_id = request.form.get(
            "department_id"
        )

        role = request.form.get(
            "role"
        )

        existing_staff = User.query.filter_by(
            staff_id=staff_id
        ).first()

        if existing_staff:

            toast(
                "Staff ID already exists",
                "warning"
            )

            return redirect(
                url_for(
                    "main.add_user"
                )
            )

        existing_username = User.query.filter_by(
            username=username
        ).first()

        if existing_username:

            toast(

                "Username already exists",
                "warning"
            )

            return redirect(

                url_for(

                    "main.add_user"

                )

            )

        existing_email = User.query.filter_by(

            email=email

        ).first()

        if existing_email:
            toast(
                "Email already in use",
                "warning"
            )
            #return redirect(
                #url_for(
                    #"main.add_user"
                #)
            #)
        ##########################################################If email already exist ###########3
        #email_pattern = r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$'

        #if not re.match(

        #    email_pattern,

        #    email
        # ):

            #toast(

                #"Invalid email address",

                #"warning"

            #)

            #return redirect(

                #url_for(

                    #"main.add_user"

                #)

            #)

        if password != confirm_password:

            toast(

                "Passwords do not match",

                "warning"

            )

            return redirect(

                url_for(

                    "main.add_user"

                )

            )

        user = User(

            staff_id=staff_id,

            username=username,

            full_name=full_name,

            email=email,

            phone=phone,

            unit_id=unit_id,

            department_id=department_id,

            role=role

        )

        user.set_password(

            password

        )

        db.session.add(

            user

        )

        try:

            db.session.commit()

        except Exception:

            db.session.rollback()

            toast(

                "Database error while creating user",

                "danger"

            )

            return redirect(

                url_for(

                    "main.add_user"

                )

            )

        log_audit(

            "Users",

            "Create User",

            f"Created user {user.username}"

        )

        toast(

            "User created successfully",

            "success"

        )

        return redirect(

            url_for(

                "main.users"

            )

        )

    return render_template(

        "users/add.html",

        units=units,

        departments=departments

    )

##################################Edit a User Route##########################################
@main.route(
    "/users/edit/<int:id>",
    methods=["GET", "POST"]
)
@login_required
def edit_user(id):

    user = User.query.get_or_404(id)

    units = Unit.query.filter_by(
        status=True
    ).all()

    departments = Department.query.filter_by(
        status=True
    ).all()

    if request.method == "POST":

        existing_email = User.query.filter(
            User.email == request.form.get("email"),
            User.id != user.id
        ).first()

        ####################################################################################
        email = request.form.get("email").strip()

        email_pattern = r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$"

        if not re.match(email_pattern, email):

            toast(
                "Please enter a valid email address",
                "warning"
            )

            return redirect(
                url_for(
                    "main.edit_user",
                    id=id
                )
            )
        ################################################################################

        if existing_email:

            toast(
                "Email already exists",
                "warning"
            )

            return redirect(
                url_for(
                    "main.edit_user",
                    id=id
                )
            )

        user.staff_id = request.form.get(
            "staff_id"
        )

        user.full_name = request.form.get(
            "full_name"
        )

        user.username = request.form.get(
            "username"
        )

        user.email = request.form.get(
            "email"
        )

        user.phone = request.form.get(
            "phone"
        )

        user.unit_id = request.form.get(
            "unit_id"
        )

        user.department_id = request.form.get(
            "department_id"
        )

        user.role = request.form.get(
            "role"
        )

        db.session.commit()

        log_audit(
            "Users",
            "Edit User",
            f"Updated {user.username}"
        )

        toast(
            "User updated successfully",
            "success"
        )

        return redirect(
            url_for("main.users")
        )

    return render_template(
        "users/edit.html",
        user=user,
        units=units,
        departments=departments
    )

@main.route("/users/toggle/<int:id>")
@login_required
def toggle_user(id):

    user = User.query.get_or_404(id)

    user.is_active_user = (
        not user.is_active_user
    )

    db.session.commit()

    log_audit(
        "Users",
        "Toggle User",
        f"{user.username} status changed"
    )

    toast(
        "User status updated",
        "info"
    )

    return redirect(
        url_for("main.users")
    )


@main.route("/logout")
@login_required
def logout():

    log_audit(
        "Authentication",
        "Logout",
        f"{current_user.username} logged out"
    )

    logout_user()

    toast(
        "Logged out successfully",
        "success"
    )

    return redirect(
        url_for("main.login")
    )

@main.route(
    "/change-password",
    methods=["GET", "POST"]
)
@login_required
def change_password():

    if request.method == "POST":

        current_password = request.form.get(
            "current_password"
        )

        new_password = request.form.get(
            "new_password"
        )

        confirm_password = request.form.get(
            "confirm_password"
        )

        if not current_user.check_password(
            current_password
        ):

            toast(
                "Current password is incorrect",
                "warning"
            )

            return redirect(
                url_for(
                    "main.change_password"
                )
            )

        if new_password != confirm_password:

            toast(
                "Passwords do not match",
                "warning"
            )

            return redirect(
                url_for(
                    "main.change_password"
                )
            )

        if len(new_password) < 8:

            toast(
                "Password must be at least 8 characters",
                "warning"
            )

            return redirect(
                url_for(
                    "main.change_password"
                )
            )

        current_user.set_password(
            new_password
        )

        current_user.must_change_password = False

        db.session.commit()

        log_audit(
            "Authentication",
            "Change Password",
            f"{current_user.username} changed password"
        )

        logout_user()

        toast(
            "Password changed successfully. Please login again.",
            "success"
        )

        return redirect(
            url_for("main.login")
        )
    return render_template(
        "auth/change_password.html"
    )

@main.route(
    "/profile",
    methods=["GET", "POST"]
)
@login_required
def profile():

    if request.method == "POST":

        current_user.full_name = request.form.get(
            "full_name"
        )

        current_user.email = request.form.get(
            "email"
        )

        current_user.phone = request.form.get(
            "phone"
        )

        file = request.files.get(
            "profile_image"
        )

        if file and file.filename:
            extension = file.filename.rsplit(
                ".",
                1
            )[1].lower()

            filename = (
                f"user_{current_user.id}.{extension}"
            )

            upload_folder = os.path.join(

                current_app.root_path,

                "static",

                "uploads",

                "profiles"

            )

            os.makedirs(
                upload_folder,
                exist_ok=True
            )

            file.save(

                os.path.join(
                    upload_folder,
                    filename
                )

            )

            current_user.profile_image = filename

        db.session.commit()

        log_audit(

            "Profile",

            "Update",

            f"{current_user.username} updated profile"

        )

        toast(

            "Profile updated successfully",

            "success"

        )

        return redirect(
            url_for(
                "main.profile"
            )
        )

    return render_template(
        "auth/profile.html"
    )
@main.route("/attendance")
@login_required
def my_attendance():

    page = request.args.get(
        "page",
        1,
        type=int
    )

    query = Attendance.query.filter_by(
        user_id=current_user.id
    )

    records = query.order_by(

        Attendance.attendance_date.desc(),

        Attendance.id.desc()

    ).paginate(

        page=page,

        per_page=20,

        error_out=False

    )

    return render_template(

        "attendance/list.html",

        records=records

    )
#####################################################Check in Route########################################################

@main.route(
    "/attendance/check-in",
    methods=["GET", "POST"]
)
@login_required
def check_in():

    today = nigeria_now().date()

    if request.method == "POST":

        existing = Attendance.query.filter_by(
            user_id=current_user.id,
            attendance_date=today
        ).first()

        if existing:

            toast(
                "You have already checked in today",
                "warning"
            )

            return redirect(
                url_for("main.my_attendance")
            )

        if not current_user.unit_id:

            toast(
                "You are not assigned to any unit",
                "warning"
            )

            return redirect(
                url_for("main.dashboard")
            )

        # ==========================================
        # GPS CAPTURE
        # ==========================================

        check_in_latitude = request.form.get(
            "latitude"
        )

        check_in_longitude = request.form.get(
            "longitude"
        )

        if not check_in_latitude or not check_in_longitude:

            toast(
                "Location is required for attendance",
                "warning"
            )

            return redirect(
                url_for("main.check_in")
            )

        # ==========================================
        # GEOFENCE VALIDATION
        # ==========================================

        active_locations = UnitLocation.query.filter_by(
            unit_id=current_user.unit_id,
            status=True
        ).all()

        if not active_locations:

            toast(
                "No active attendance location configured for your unit",
                "danger"
            )

            return redirect(
                url_for("main.my_attendance")
            )

        allowed = False

        matched_location = None

        matched_distance = None

        for location in active_locations:

            distance = distance_in_meters(

                check_in_latitude,
                check_in_longitude,

                location.latitude,
                location.longitude
            )

            if distance <= location.radius:

                allowed = True

                matched_location = location

                matched_distance = distance

                break

        if not allowed:

            toast(
                "You are outside all approved attendance locations",
                "danger"
            )

            return redirect(
                url_for("main.my_attendance")
            )

        # ==========================================
        # SELFIE CAPTURE
        # ==========================================

        photo_data = request.form.get(
            "photo"
        )

        filename = None

        if photo_data:

            upload_folder = os.path.join(
                "app",
                "static",
                "uploads",
                "attendance"
            )

            os.makedirs(
                upload_folder,
                exist_ok=True
            )

            filename = (
                f"checkin_"
                f"{current_user.id}_"
                f"{today}.jpg"
            )

            filepath = os.path.join(
                upload_folder,
                filename
            )

            try:

                header, encoded = photo_data.split(
                    ",",
                    1
                )

                with open(
                    filepath,
                    "wb"
                ) as f:

                    f.write(
                        base64.b64decode(
                            encoded
                        )
                    )

            except Exception as e:

                toast(
                    f"Photo save failed: {e}",
                    "danger"
                )

                return redirect(
                    url_for(
                        "main.check_in"
                    )
                )

        # ==========================================
        # SAVE ATTENDANCE
        # ==========================================

        attendance = Attendance(

            user_id=current_user.id,

            unit_id=current_user.unit_id,

            location_id=matched_location.id,

            attendance_date=today,

            check_in=nigeria_now(),

            check_in_photo=filename,

            check_in_latitude=check_in_latitude,

            check_in_longitude=check_in_longitude,

            status="Present"
        )

        db.session.add(
            attendance
        )

        db.session.commit()

        # ==========================================
        # AUDIT LOG
        # ==========================================

        log_audit(
            "Attendance",
            "Check In",
            f"{current_user.username} checked in "
            f"{round(matched_distance)}m from "
            f"{matched_location.name}"
        )

        toast(
            "Check in successful",
            "success"
        )

        return redirect(
            url_for("main.my_attendance")
        )

    return render_template(
        "attendance/check_in.html"
    )

@main.route(
    "/attendance/check-out",
    methods=["GET", "POST"]
)

##########################################################################Check out Route#####################################################################
##############################################################################################################################################
@main.route(
    "/attendance/check-out",
    methods=["GET", "POST"]
)
@login_required
def check_out():

    today = nigeria_now().date()

    attendance = Attendance.query.filter_by(
        user_id=current_user.id,
        attendance_date=today
    ).order_by(
        Attendance.id.desc()
    ).first()

    if not attendance:

        toast(
            "You must check in first",
            "warning"
        )

        return redirect(
            url_for("main.my_attendance")
        )

    if request.method == "POST":

        if attendance.check_out:

            toast(
                "Already checked out",
                "warning"
            )

            return redirect(
                url_for("main.my_attendance")
            )

        # ==========================================
        # GPS CAPTURE
        # ==========================================

        check_out_latitude = request.form.get(
            "latitude"
        )

        check_out_longitude = request.form.get(
            "longitude"
        )

        if not check_out_latitude or not check_out_longitude:

            toast(
                "Location is required for check out. Kindly allow location access and try again.",
                "warning"
            )

            return redirect(
                url_for("main.check_out")
            )

        # ==========================================
        # GEOFENCE VALIDATION
        # ==========================================

        active_locations = UnitLocation.query.filter_by(
            unit_id=current_user.unit_id,
            status=True
        ).all()

        if not active_locations:

            toast(
                "No active attendance location configured for your unit",
                "danger"
            )

            return redirect(
                url_for("main.my_attendance")
            )

        allowed = False

        matched_location = None

        matched_distance = None

        for location in active_locations:

            distance = distance_in_meters(

                check_out_latitude,
                check_out_longitude,

                location.latitude,
                location.longitude
            )

            if distance <= location.radius:

                allowed = True

                matched_location = location

                matched_distance = distance

                break

        if not allowed:

            toast(
                "You are outside all approved attendance locations",
                "danger"
            )

            return redirect(
                url_for("main.my_attendance")
            )

        # ==========================================
        # SELFIE CAPTURE
        # ==========================================

        photo_data = request.form.get(
            "photo"
        )

        filename = None

        if photo_data:

            upload_folder = os.path.join(
                "app",
                "static",
                "uploads",
                "attendance"
            )

            os.makedirs(
                upload_folder,
                exist_ok=True
            )

            filename = (
                f"checkout_"
                f"{current_user.id}_"
                f"{today}.jpg"
            )

            filepath = os.path.join(
                upload_folder,
                filename
            )

            try:

                header, encoded = photo_data.split(
                    ",",
                    1
                )

                with open(
                    filepath,
                    "wb"
                ) as f:

                    f.write(
                        base64.b64decode(
                            encoded
                        )
                    )

            except Exception as e:

                toast(
                    f"Photo save failed: {e}",
                    "danger"
                )

                return redirect(
                    url_for(
                        "main.check_out"
                    )
                )

        # ==========================================
        # SAVE CHECK OUT
        # ==========================================

        attendance.check_out = nigeria_now()

        attendance.check_out_photo = filename

        attendance.check_out_latitude = (
            check_out_latitude
        )

        attendance.check_out_longitude = (
            check_out_longitude
        )

        db.session.commit()

        # ==========================================
        # AUDIT LOG
        # ==========================================

        log_audit(
            "Attendance",
            "Check Out",
            f"{current_user.username} checked out "
            f"{round(matched_distance)}m from "
            f"{matched_location.name}"
        )

        toast(
            "Check out successful",
            "success"
        )

        return redirect(
            url_for("main.my_attendance")
        )

    return render_template(
        "attendance/check_out.html"
    )

#######################Reusable Filter Block for Attendance Report Route############################################################
###############################################################################################################################

def get_attendance_query():

    query = Attendance.query

    start_date = request.args.get(
        "start_date"
    )

    end_date = request.args.get(
        "end_date"
    )

    unit_id = request.args.get(
        "unit_id"
    )

    user_id = request.args.get(
        "user_id"
    )

    if start_date:

        query = query.filter(
            Attendance.attendance_date >= start_date
        )

    if end_date:

        query = query.filter(
            Attendance.attendance_date <= end_date
        )

    if unit_id:

        query = query.filter(
            Attendance.unit_id == unit_id
        )

    if user_id:

        query = query.filter(
            Attendance.user_id == user_id
        )

    return query


###################################################################################################333333333333333
#############################Attendance Report Route########################################################################################
@main.route(
    "/reports/attendance",
    methods=["GET"]
)
@login_required
def attendance_report():

    page = request.args.get(
        "page",
        1,
        type=int
    )

    start_date = request.values.get(
        "start_date"
    )

    end_date = request.values.get(
        "end_date"
    )

    unit_id = request.values.get(
        "unit_id"
    )

    user_id = request.values.get(
        "user_id"
    )

    units = Unit.query.filter_by(
        status=True
    ).all()

    users = User.query.filter_by(
        is_active_user=True
    ).all()

    query = Attendance.query

    if start_date:

        query = query.filter(
            Attendance.attendance_date >= start_date
        )

    if end_date:

        query = query.filter(
            Attendance.attendance_date <= end_date
        )

    if unit_id:

        query = query.filter(
            Attendance.unit_id == unit_id
        )

    if user_id:

        query = query.filter(
            Attendance.user_id == user_id
        )

    records = query.order_by(

        Attendance.attendance_date.desc(),

        Attendance.id.desc()

    ).paginate(

        page=page,

        per_page=20,

        error_out=False

    )

    if start_date or end_date or unit_id or user_id:

        log_audit(
            "Reports",
            "Attendance Report",
            "Attendance report generated"
        )

    return render_template(

        "reports/attendance.html",

        records=records,

        start_date=start_date,

        end_date=end_date,

        unit_id=unit_id,

        user_id=user_id,

        units=units,

        users=users
    )




@main.route("/audit-logs",methods=["GET", "POST"])
@login_required
@superadmin_required
def audit_logs():

    users = User.query.order_by(
        User.full_name
    ).all()

    modules = db.session.query(
        AuditLog.module
    ).distinct().all()

    query = AuditLog.query
    total_logs = AuditLog.query.count()

    today_logs = AuditLog.query.filter(
        db.func.date(
            AuditLog.created_at
        ) == nigeria_now().date()
    ).count()

    total_users = db.session.query(
        AuditLog.username
    ).distinct().count()

    total_modules = db.session.query(
        AuditLog.module
    ).distinct().count()

    start_date = None
    end_date = None
    user_id = None
    module = None

    if request.method == "POST":

        start_date = request.form.get(
            "start_date"
        )

        end_date = request.form.get(
            "end_date"
        )

        user_id = request.form.get(
            "user_id"
        )

        module = request.form.get(
            "module"
        )

        if start_date:

            query = query.filter(
                db.func.date(
                    AuditLog.created_at
                ) >= start_date
            )

        if end_date:

            query = query.filter(
                db.func.date(
                    AuditLog.created_at
                ) <= end_date
            )

        if user_id:

            query = query.filter(
                AuditLog.user_id == int(user_id)
            )

        if module:

            query = query.filter(
                AuditLog.module == module
            )

    page = request.args.get(
        "page",
        1,
        type=int
    )

    logs = query.order_by(
        AuditLog.id.desc()
    ).paginate(
        page=page,
        per_page=50
    )

    return render_template(
        "audit_logs/list.html",
        logs=logs,
        users=users,
        modules=modules,
        start_date=start_date,
        end_date=end_date,
        selected_user=user_id,
        total_logs=total_logs,
        today_logs=today_logs,
        total_users=total_users,
        total_modules=total_modules,
        selected_module=module
    )

@main.route(
    "/audit-logs/excel"
)
@login_required
@superadmin_required
def export_audit_excel():

    logs = AuditLog.query.order_by(
        AuditLog.id.desc()
    ).all()

    data = []

    for log in logs:

        data.append({

            "Date Time":
            log.created_at,

            "User":
            log.username,

            "Module":
            log.module,

            "Action":
            log.action,

            "Details":
            log.details,

            "IP Address":
            log.ip_address
        })

    df = pd.DataFrame(data)

    output = BytesIO()

    df.to_excel(
        output,
        index=False
    )

    output.seek(0)

    return send_file(

        output,

        download_name=
        "audit_logs.xlsx",

        as_attachment=True
    )
@main.route(
    "/audit-logs/pdf"
)
@login_required
@admin_required
def export_audit_pdf():

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer
    )

    data = [[

        "Date",

        "User",

        "Module",

        "Action"

    ]]

    logs = AuditLog.query.order_by(
        AuditLog.id.desc()
    ).all()

    for log in logs:

        data.append([

            str(log.created_at),

            log.username or "",

            log.module or "",

            log.action or ""

        ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightblue
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.black
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            )

        ])
    )

    doc.build(
        [table]
    )

    buffer.seek(0)

    return send_file(

        buffer,

        as_attachment=True,

        download_name=
        "audit_logs.pdf",

        mimetype=
        "application/pdf"
    )

@main.route(
    "/audit-logs/print"
)
@login_required
@admin_required
def print_audit_logs():

    logs = AuditLog.query.order_by(
        AuditLog.id.desc()
    ).all()

    return render_template(

        "audit_logs/print.html",

        logs=logs,

        generated_at=datetime.now().strftime(
            "%d-%b-%Y %H:%M"
        )
    )



@main.route("/reports/attendance/print")
@login_required
@admin_required
def print_attendance_report():

    records = get_attendance_query().order_by(
        Attendance.attendance_date.desc()
        ).all()

    generated_at = datetime.now().strftime(
        "%d-%b-%Y %I:%M %p"
    )
    return render_template(
        "reports/attendance_print.html",
        records=records,
        generated_at=generated_at
    )

@main.route(
    "/reports/attendance/excel"
)
@login_required
@admin_required
def export_attendance_excel():

    records = get_attendance_query().order_by(
        Attendance.attendance_date.desc()
        ).all()

    data = []

    for record in records:

        data.append({

            "Date":
            record.attendance_date,

            "Staff":
            record.user.full_name,

            "Unit":
            record.unit.name,

            "Location":
            record.location.name
            if record.location
            else "",

            "Check In":
            record.check_in,

            "Check Out":
            record.check_out,

            "Status":
            record.status,

            "Check In Latitude":
            record.check_in_latitude,

            "Check In Longitude":
            record.check_in_longitude,

            "Check Out Latitude":
            record.check_out_latitude,

            "Check Out Longitude":
            record.check_out_longitude
        })

    df = pd.DataFrame(data)

    output = BytesIO()

    df.to_excel(
        output,
        index=False
    )

    output.seek(0)

    return send_file(

        output,

        download_name=
        "attendance_report.xlsx",

        as_attachment=True
    )

@main.route(
    "/reports/attendance/pdf"
)
@login_required
@admin_required
def export_attendance_pdf():

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer
    )

    data = [[

        "Date",

        "Staff",

        "Unit",

        "Location",

        "Status"

    ]]

    records = get_attendance_query().order_by(
        Attendance.attendance_date.desc()
    ).all()

    for record in records:

        data.append([

            str(
                record.attendance_date
            ),

            record.user.full_name,

            record.unit.name,

            record.location.name
            if record.location
            else "",

            record.status

        ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightblue
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.black
            )
        ])
    )

    doc.build(
        [table]
    )

    buffer.seek(0)

    return send_file(

        buffer,

        download_name=
        "attendance_report.pdf",

        as_attachment=True,

        mimetype=
        "application/pdf"
    )

@main.route(
    "/field-attendance/checkpoint",
    methods=["GET", "POST"]
)
@login_required
def field_checkpoint():

    today = nigeria_now().date()

    current_time = nigeria_now().time()

    current_punch = get_current_punch(
        current_time
    )

    if not current_punch:

        toast(
            "No active attendance window currently, its late.",
            "warning"
        )

        return redirect(
            url_for(
                "main.dashboard"
            )
        )

    existing = FieldAttendance.query.filter_by(

        user_id=current_user.id,

        attendance_date=today,

        checkpoint_number=current_punch

    ).first()

    if existing:

        toast(
            f"Punch {current_punch} already recorded",
            "warning"
        )

        return redirect(
            url_for(
                "main.dashboard"
            )
        )

    if request.method == "POST":

        photo_data = request.form.get(
            "photo"
        )

        latitude = request.form.get(
            "latitude"
        )

        longitude = request.form.get(
            "longitude"
        )

        remarks = request.form.get(
            "remarks"
        )

        if not latitude or not longitude:

            toast(
                "Location required",
                "danger"
            )

            return redirect(
                url_for(
                    "main.field_checkpoint"
                )
            )

        active_locations = UnitLocation.query.filter_by(

            unit_id=current_user.unit_id,

            status=True

        ).all()

        if not active_locations:

            toast(
                "No attendance location configured",
                "danger"
            )

            return redirect(
                url_for(
                    "main.dashboard"
                )
            )

        allowed = False

        matched_location = None

        matched_distance = None

        for location in active_locations:

            distance = distance_in_meters(

                float(latitude),

                float(longitude),

                location.latitude,

                location.longitude

            )

            if distance <= location.radius:

                allowed = True

                matched_location = location

                matched_distance = distance

                break

        if not allowed:

            toast(
                "You are outside all approved attendance locations",
                "danger"
            )

            return redirect(
                url_for(
                    "main.field_checkpoint"
                )
            )
        
        if not photo_data:
            toast(
                "Selfie is required",
                "danger"
                )
            return redirect(
                    url_for(
                        "main.field_checkpoint"
                    )
                )

        upload_folder = os.path.join(

            "app",

            "static",

            "uploads",

            "field_attendance"

        )

        os.makedirs(

            upload_folder,

            exist_ok=True

        )

        filename = (

            f"field_"

            f"{current_user.id}_"

            f"{today}_"

            f"{current_punch}.jpg"

        )

        filepath = os.path.join(

            upload_folder,

            filename

        )

        header, encoded = photo_data.split(
            ",",
            1
        )

        with open(
            filepath,
            "wb"
        ) as f:

            f.write(
                base64.b64decode(
                    encoded
                )
            )
        attendance = FieldAttendance(

            user_id=current_user.id,

            unit_id=current_user.unit_id,

            location_id=matched_location.id,

            attendance_date=today,

            checkpoint_number=current_punch,

            attendance_time=nigeria_now(),

            photo=filename,

            latitude=float(latitude),

            longitude=float(longitude),

            distance=matched_distance,

            remarks=remarks

        )

        db.session.add(
            attendance
        )

        db.session.commit()

        log_audit(

            "Field Attendance",

            f"Punch {current_punch}",

            f"{current_user.username} completed Punch {current_punch} "
            f"at {matched_location.name} "
            f"({round(matched_distance,2)}m)"
        )

        toast(

            f"Punch {current_punch} recorded successfully",

            "success"

        )

        return redirect(
            url_for(
                "main.dashboard"
            )
        )

    today_count = FieldAttendance.query.filter_by(

        user_id=current_user.id,

        attendance_date=today

    ).count()

    return render_template(

        "field_attendance/checkpoint.html",

        current_punch=current_punch,

        completed=today_count,

        remaining=8 - today_count
    )

@main.route(
    "/reports/field-attendance",
    methods=["GET", "POST"]
)
@login_required
@admin_required
def field_attendance_report():

    page = request.args.get(
        "page",
        1,
        type=int
    )

    query = FieldAttendance.query

    start_date = request.values.get(
        "start_date"
    )

    end_date = request.values.get(
        "end_date"
    )

    user_id = request.values.get(
        "user_id"
    )
    unit_id = request.values.get(
        "unit_id"
    )
    

    if start_date:

        query = query.filter(
            FieldAttendance.attendance_date >= start_date
        )

    if end_date:

        query = query.filter(
            FieldAttendance.attendance_date <= end_date
        )

    if user_id:

        query = query.filter(
            FieldAttendance.user_id == user_id
        )

    # ==========================
    # STATISTICS
    # ==========================

    total_punches = query.count()

    today_punches = FieldAttendance.query.filter_by(
        attendance_date=nigeria_now().date()
    ).count()

    active_staff = User.query.filter_by(
        is_active_user=True
    ).count()

    compliance = round(

        (today_punches / max(1, active_staff * 8)) * 100,

        1

    )

    # ==========================
    # RECORDS
    # ==========================
    if unit_id:

        query = query.filter(
            FieldAttendance.unit_id == unit_id
        )

    records = query.order_by(

        FieldAttendance.attendance_date.desc(),

        FieldAttendance.attendance_time.desc()

    ).paginate(

        page=page,

        per_page=20,

        error_out=False
    )

    users = User.query.order_by(
        User.full_name
    ).all()

    units = Unit.query.filter_by(
        status=True).order_by(
            Unit.name).all()

    return render_template(

        "reports/field_attendance.html",

        records=records,

        users=users,

        start_date=start_date,

        end_date=end_date,

        user_id=user_id,
        units=units,
        unit_id=unit_id,

        total_punches=total_punches,

        today_punches=today_punches,

        active_staff=active_staff,

        compliance=compliance
    )
@main.route(
    "/reports/field-attendance/print"
)
@login_required
@admin_required
def print_field_attendance():

    query = FieldAttendance.query

    start_date = request.args.get(
        "start_date"
    )

    end_date = request.args.get(
        "end_date"
    )

    user_id = request.args.get(
        "user_id"
    )

    if start_date:

        query = query.filter(
            FieldAttendance.attendance_date >= start_date
        )

    if end_date:

        query = query.filter(
            FieldAttendance.attendance_date <= end_date
        )

    if user_id:

        query = query.filter(
            FieldAttendance.user_id == user_id
        )

    records = query.order_by(

        FieldAttendance.attendance_date.desc(),

        FieldAttendance.attendance_time.desc()

    ).all()

    return render_template(

        "reports/field_attendance_print.html",

        records=records,

        generated_at=nigeria_now()
    )

@main.route(
    "/reports/field-attendance/compliance"
)
@login_required
@admin_required
def field_attendance_compliance():

    selected_date = request.args.get(
        "report_date"
    )
    user_id = request.args.get(
        "user_id"
    )

    unit_id = request.args.get(
        "unit_id"
    )

    if selected_date:

        report_date = datetime.strptime(
            selected_date,
            "%Y-%m-%d"
        ).date()

    else:

        report_date = nigeria_now().date()

    
    staff_query = User.query.filter_by(
        is_active_user=True
    )

    if unit_id:

        staff_query = staff_query.filter(
            User.unit_id == unit_id
        )

    if user_id:

        staff_query = staff_query.filter(
            User.id == user_id
        )

    staff = staff_query.order_by(
        User.full_name
    ).all()

    compliance_data = []

    for user in staff:

        punches = FieldAttendance.query.filter_by(

            user_id=user.id,

            attendance_date=report_date

        ).all()

        completed = [

            punch.checkpoint_number

            for punch in punches
        ]

        compliance = round(

            (len(completed) / 8) * 100,

            1
        )

        compliance_data.append({

            "user": user,

            "completed": completed,

            "compliance": compliance

        })
    units = Unit.query.filter_by(
        status=True
    ).order_by(
        Unit.name
    ).all()

    users = User.query.filter_by(
        is_active_user=True
    ).order_by(
        User.full_name
    ).all()


    return render_template(

        "reports/field_attendance_compliance.html",

        compliance_data=compliance_data,

        report_date=report_date,

        users=users,

        units=units,

        user_id=user_id,

        unit_id=unit_id
    )

#######################################################################################
######################field attendance Report Export###################################
######################################################################################

@main.route(
    "/reports/field-attendance/excel"
)
@login_required
@admin_required
def export_field_attendance_excel():

    query = FieldAttendance.query

    start_date = request.args.get(
        "start_date"
    )

    end_date = request.args.get(
        "end_date"
    )

    user_id = request.args.get(
        "user_id"
    )

    if start_date:

        query = query.filter(
            FieldAttendance.attendance_date >= start_date
        )

    if end_date:

        query = query.filter(
            FieldAttendance.attendance_date <= end_date
        )

    if user_id:

        query = query.filter(
            FieldAttendance.user_id == user_id
        )

    records = query.order_by(

        FieldAttendance.attendance_date.desc(),

        FieldAttendance.attendance_time.desc()

    ).all()

    data = []

    for record in records:

        data.append({

            "Date":
            record.attendance_date,

            "Staff":
            record.user.full_name,

            "Punch":
            record.checkpoint_number,

            "Location":
            record.location.name,

            "Time":
            record.attendance_time.strftime(
                "%I:%M %p"
            ),

            "Distance (m)":
            round(
                record.distance,
                2
            ),

            "Remarks":
            record.remarks
        })

    df = pd.DataFrame(
        data
    )

    output = BytesIO()

    df.to_excel(

        output,

        index=False
    )

    output.seek(0)

    return send_file(

        output,

        download_name=
        "field_attendance.xlsx",

        as_attachment=True
    )
@main.route(
    "/reports/field-attendance/pdf"
)
@login_required
@admin_required
def export_field_attendance_pdf():

    query = FieldAttendance.query

    start_date = request.args.get(
        "start_date"
    )

    end_date = request.args.get(
        "end_date"
    )

    user_id = request.args.get(
        "user_id"
    )

    if start_date:

        query = query.filter(
            FieldAttendance.attendance_date >= start_date
        )

    if end_date:

        query = query.filter(
            FieldAttendance.attendance_date <= end_date
        )

    if user_id:

        query = query.filter(
            FieldAttendance.user_id == user_id
        )

    records = query.order_by(

        FieldAttendance.attendance_date.desc(),

        FieldAttendance.attendance_time.desc()

    ).all()

    output = BytesIO()

    doc = SimpleDocTemplate(
        output
    )

    data = [[

        "Date",

        "Staff",

        "Punch",

        "Location",

        "Time",

        "Distance",

        "Remarks"

    ]]

    for record in records:

        data.append([

            str(
                record.attendance_date
            ),

            record.user.full_name,

            f"Punch {record.checkpoint_number}",

            record.location.name,

            record.attendance_time.strftime(
                "%I:%M %p"
            ),

            f"{round(record.distance,2)} m",

            record.remarks or ""
        ])

    table = Table(
        data
    )

    table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0,0),
                (-1,0),
                colors.lightblue
            ),

            (
                "GRID",
                (0,0),
                (-1,-1),
                1,
                colors.black
            ),

            (
                "FONTSIZE",
                (0,0),
                (-1,-1),
                8
            )

        ])
    )

    doc.build(
        [table]
    )

    output.seek(0)

    return send_file(

        output,

        download_name=
        "field_attendance.pdf",

        as_attachment=True,

        mimetype=
        "application/pdf"
    )

#########################################################################################################
@main.route(
    "/reports/field_attend_compl",
    methods=["GET", "POST"]
)
@login_required
@admin_required
def field_attend_compl():

    start_date = request.values.get(
        "start_date"
    )

    end_date = request.values.get(
        "end_date"
    )

    user_id = request.values.get(
        "user_id"
    )

    unit_id = request.values.get(
        "unit_id"
    )

    # ==========================
    # ATTENDANCE QUERY
    # ==========================

    attendance_query = FieldAttendance.query

    if start_date:

        attendance_query = attendance_query.filter(
            FieldAttendance.attendance_date >= start_date
        )

    if end_date:

        attendance_query = attendance_query.filter(
            FieldAttendance.attendance_date <= end_date
        )

    if user_id:

        attendance_query = attendance_query.filter(
            FieldAttendance.user_id == user_id
        )

    if unit_id:

        attendance_query = attendance_query.filter(
            FieldAttendance.unit_id == unit_id
        )

    # ==========================
    # SUMMARY CARDS
    # ==========================

    actual_punches = attendance_query.count()

    staff_query = User.query.filter_by(
        is_active_user=True
    )

    if unit_id:

        staff_query = staff_query.filter(
            User.unit_id == unit_id
        )

    if user_id:

        staff_query = staff_query.filter(
            User.id == user_id
        )

    active_staff = staff_query.count()
    number_of_days = 1

    if start_date and end_date:

        start_dt = datetime.strptime(
            start_date,
            "%Y-%m-%d"
        ).date()

        end_dt = datetime.strptime(
            end_date,
            "%Y-%m-%d"
        ).date()

        number_of_days = (
            end_dt - start_dt
        ).days + 1

    expected_punches = (active_staff * 8 * number_of_days)

    compliance = round(

        (
            actual_punches /
            max(1, expected_punches)
        ) * 100,

        1

    )
    
    # ==========================
    # STAFF SUMMARY REPORT
    # ==========================

    report_data = []

    staff_list = staff_query.order_by(
        User.full_name
    ).all()

    for staff in staff_list:

        punches_query = FieldAttendance.query.filter(
            FieldAttendance.user_id == staff.id
        )

        if start_date:

            punches_query = punches_query.filter(
                FieldAttendance.attendance_date >= start_date
            )

        if end_date:

            punches_query = punches_query.filter(
                FieldAttendance.attendance_date <= end_date
            )

        if unit_id:

            punches_query = punches_query.filter(
                FieldAttendance.unit_id == unit_id
            )

        total_staff_punches = punches_query.count()
        staff_expected = ( 8 * number_of_days)
        staff_compliance = round(
            (
                total_staff_punches /
                max(1, staff_expected)
            ) * 100,

            1

        )
    ############################################3

        report_data.append({

            "staff_name": staff.full_name,

            "unit_name": (
                staff.unit.name
                if hasattr(staff, "unit")
                and staff.unit
                else ""
            ),

            "total_punches": total_staff_punches,
            "expected_punches": staff_expected,
            "compliance": staff_compliance

        })

    # ==========================
    # FILTER DROPDOWNS
    # ==========================

    users = User.query.order_by(
        User.full_name
    ).all()

    units = Unit.query.filter_by(
        status=True
    ).order_by(
        Unit.name
    ).all()
    selected_unit = "All Units"

    if unit_id:

        selected_unit_obj = Unit.query.get(unit_id)

        if selected_unit_obj:

            selected_unit = selected_unit_obj.name

    return render_template(

        "reports/field_attendance_compliance_page.html",

        users=users,

        units=units,

        start_date=start_date,

        end_date=end_date,

        user_id=user_id,

        unit_id=unit_id,

        total_punches=actual_punches,

        expected_punches=expected_punches,

        active_staff=active_staff,

        compliance=compliance,
        selected_unit=selected_unit,
        report_data=report_data

    )


@main.route(
    "/reports/field-attendance/compliance/details"
)
@login_required
@admin_required
def field_attendance_compliance_details():

    start_date = request.args.get(
        "start_date"
    )

    end_date = request.args.get(
        "end_date"
    )

    unit_id = request.args.get(
        "unit_id"
    )

    user_id = request.args.get(
        "user_id"
    )

    query = FieldAttendance.query

    if start_date:

        query = query.filter(
            FieldAttendance.attendance_date >= start_date
        )

    if end_date:

        query = query.filter(
            FieldAttendance.attendance_date <= end_date
        )

    if unit_id:

        query = query.filter(
            FieldAttendance.unit_id == unit_id
        )

    if user_id:

        query = query.filter(
            FieldAttendance.user_id == user_id
        )

    attendance_records = query.order_by(
        FieldAttendance.attendance_date.desc(),
        FieldAttendance.attendance_time.desc()
    ).all()

    return render_template(
        "reports/field_attendance_compliance_details.html",
        attendance_records=attendance_records
    )

@main.route(
    "/reports/field_attend_compl/excel"
)
@login_required
@admin_required
def field_attend_compl_excel():

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    user_id = request.args.get("user_id")
    unit_id = request.args.get("unit_id")

    staff_query = User.query.filter_by(
        is_active_user=True
    )

    if unit_id:

        staff_query = staff_query.filter(
            User.unit_id == unit_id
        )

    if user_id:

        staff_query = staff_query.filter(
            User.id == user_id
        )

    report_data = []

    for staff in staff_query.all():

        punches_query = FieldAttendance.query.filter(
            FieldAttendance.user_id == staff.id
        )

        if start_date:

            punches_query = punches_query.filter(
                FieldAttendance.attendance_date >= start_date
            )

        if end_date:

            punches_query = punches_query.filter(
                FieldAttendance.attendance_date <= end_date
            )

        total_punches = punches_query.count()

        report_data.append({

            "Staff": staff.full_name,

            "Unit": staff.unit.name
            if staff.unit else "",

            "Total Punches": total_punches,

            "Compliance %": round(
                (total_punches / 8) * 100,
                1
            )

        })

    df = pd.DataFrame(report_data)

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name="Compliance Report"
        )

    output.seek(0)

    return send_file(

        output,

        as_attachment=True,

        download_name="field_attendance_compliance.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@main.route(
    "/reports/field_attend_compl/pdf"
)
@login_required
@admin_required
def field_attend_compl_pdf():

    start_date = request.args.get(
        "start_date"
    )

    end_date = request.args.get(
        "end_date"
    )

    user_id = request.args.get(
        "user_id"
    )

    unit_id = request.args.get(
        "unit_id"
    )

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(

        Paragraph(

            "Field Attendance Compliance Report",

            styles["Title"]

        )

    )

    elements.append(
        Spacer(1, 12)
    )

    data = [

        [

            "Staff",

            "Unit",

            "Punches",

            "Compliance %"

        ]

    ]

    staff_query = User.query.filter_by(
        is_active_user=True
    )

    if unit_id:

        staff_query = staff_query.filter(
            User.unit_id == unit_id
        )

    if user_id:

        staff_query = staff_query.filter(
            User.id == user_id
        )

    for staff in staff_query.all():

        punches_query = FieldAttendance.query.filter(
            FieldAttendance.user_id == staff.id
        )

        if start_date:

            punches_query = punches_query.filter(
                FieldAttendance.attendance_date >= start_date
            )

        if end_date:

            punches_query = punches_query.filter(
                FieldAttendance.attendance_date <= end_date
            )

        punches = punches_query.count()

        compliance = round(
            (punches / 8) * 100,
            1
        )

        data.append([

            staff.full_name,

            staff.unit.name
            if staff.unit else "",

            punches,

            f"{compliance}%"

        ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("GRID", (0, 0), (-1, -1), 1, colors.black)

        ])

    )

    elements.append(table)

    doc.build(elements)

    buffer.seek(0)

    return send_file(

        buffer,

        as_attachment=True,

        download_name="field_attendance_compliance.pdf",

        mimetype="application/pdf"

    )

@main.route(
    "/users/reset-password/<int:id>",
    methods=["GET", "POST"]
)
@login_required
@admin_required
def reset_user_password(id):

    user = User.query.get_or_404(id)

    if request.method == "POST":

        new_password = request.form.get(
            "password"
        )

        confirm_password = request.form.get(
            "confirm_password"
        )

        if new_password != confirm_password:

            toast(
                "Passwords do not match",
                "danger"
            )

            return redirect(
                url_for(
                    "main.reset_user_password",
                    id=id
                )
            )

        user.set_password(
            new_password
        )

        user.must_change_password = True

        db.session.commit()

        log_audit(

            "User Management",

            "Reset Password",

            f"Password reset for {user.username}"

        )

        toast(

            "Password reset successfully",

            "success"

        )

        return redirect(
            url_for(
                "main.users"
            )
        )

    return render_template(

        "users/reset_password.html",

        user=user

    )

@main.route(
    "/attendance-schedule-config",
    methods=["GET", "POST"]
)
@login_required
@admin_required
def attendance_schedule_config():

    config = AttendanceScheduleConfig.query.first()

    units = Unit.query.filter_by(
        status=True
    ).all()

    if request.method == "POST":

        unit_id = request.form.get(
            "unit_id"
        )

        total_checkpoints = int(
            request.form.get(
                "total_checkpoints"
            )
        )

        start_time = datetime.strptime(

            request.form.get(
                "start_time"
            ),

            "%H:%M"

        ).time()

        checkpoint_duration = int(

            request.form.get(
                "checkpoint_duration"
            )
        )

        interval_minutes = int(

            request.form.get(
                "interval_minutes"
            )
        )

        if config:

            config.unit_id = unit_id

            config.total_checkpoints = (
                total_checkpoints
            )

            config.start_time = (
                start_time
            )

            config.checkpoint_duration = (
                checkpoint_duration
            )

            config.interval_minutes = (
                interval_minutes
            )

        else:

            config = AttendanceScheduleConfig(

                unit_id=unit_id,

                total_checkpoints=(
                    total_checkpoints
                ),

                start_time=start_time,

                checkpoint_duration=(
                    checkpoint_duration
                ),

                interval_minutes=(
                    interval_minutes
                )
            )

            db.session.add(
                config
            )

        db.session.commit()

        toast(
            "Configuration saved",
            "success"
        )

        return redirect(
            url_for(
                "main.attendance_schedule_config"
            )
        )

    return render_template(

        "attendance_schedule/config.html",

        config=config,

        units=units
    )

@main.route(
    "/generate-attendance-schedule"
)
@login_required
@admin_required
def generate_attendance_schedule():

    config = AttendanceScheduleConfig.query.first()

    if not config:

        toast(
            "Please create a schedule configuration first",
            "warning"
        )

        return redirect(
            url_for(
                "main.attendance_schedule_config"
            )
        )
    start_datetime = datetime.combine(
        date.today(),
        config.start_time
    )

    AttendanceSchedule.query.filter_by(
        unit_id=config.unit_id
    ).delete()

    for checkpoint in range(
        1,
        config.total_checkpoints + 1
    ):

        open_time = start_datetime.time()

        close_time = (
            start_datetime +
            timedelta(
                minutes=config.checkpoint_duration
            )
        ).time()

        schedule = AttendanceSchedule(

            unit_id=config.unit_id,

            checkpoint_number=checkpoint,

            open_time=open_time,

            close_time=close_time,

            is_active=True
        )

        db.session.add(
            schedule
        )

        start_datetime = (
            start_datetime +
            timedelta(
                minutes=config.interval_minutes
            )
        )

    db.session.commit()

    toast(
        "Attendance schedule generated successfully",
        "success"
    )

    return redirect(
        url_for(
            "main.attendance_schedule_list"
        )
    )

@main.route("/attendance-schedules")
@login_required
@admin_required
def attendance_schedule_list():

    unit_id = request.args.get(
        "unit_id"
    )

    units = Unit.query.filter_by(
        status=True
    ).order_by(
        Unit.name
    ).all()

    query = AttendanceSchedule.query

    if unit_id:

        query = query.filter(
            AttendanceSchedule.unit_id == unit_id
        )

    schedules = query.order_by(
        AttendanceSchedule.unit_id,
        AttendanceSchedule.checkpoint_number
    ).all()

    return render_template(

        "attendance_schedule/list.html",

        schedules=schedules,

        units=units,

        unit_id=unit_id
    )

@main.route(
    "/attendance-schedule/<int:id>/toggle"
)
@login_required
@admin_required
def toggle_attendance_schedule(id):

    schedule = AttendanceSchedule.query.get_or_404(
        id
    )

    schedule.is_active = (
        not schedule.is_active
    )

    db.session.commit()

    toast(
        "Schedule updated successfully",
        "success"
    )

    return redirect(
        url_for(
            "main.attendance_schedule_list"
        )
    )

@main.route(
    "/attendance-schedule/<int:id>/delete"
)
@login_required
@admin_required
def delete_attendance_schedule(id):

    schedule = AttendanceSchedule.query.get_or_404(
        id
    )

    db.session.delete(
        schedule
    )

    db.session.commit()

    toast(
        "Checkpoint deleted successfully",
        "success"
    )

    return redirect(
        url_for(
            "main.attendance_schedule_list"
        )
    )

@main.route(
    "/attendance-schedule/<int:id>/edit",
    methods=["GET", "POST"]
)
@login_required
@admin_required
def edit_attendance_schedule(id):

    schedule = AttendanceSchedule.query.get_or_404(
        id
    )

    if request.method == "POST":

        schedule.open_time = datetime.strptime(

            request.form.get(
                "open_time"
            ),

            "%H:%M"

        ).time()

        schedule.close_time = datetime.strptime(

            request.form.get(
                "close_time"
            ),

            "%H:%M"

        ).time()

        db.session.commit()

        toast(
            "Checkpoint updated successfully",
            "success"
        )

        return redirect(
            url_for(
                "main.attendance_schedule_list"
            )
        )

    return render_template(

        "attendance_schedule/edit.html",

        schedule=schedule
    )

###################################################################################
###################333Building Field Attendance V2#################################
###################################################################################
@main.route(
    "/field-attendance-v2/checkpoint",
    methods=["GET"]
)
@login_required
def dynamic_field_checkpoint():

    schedule = get_current_checkpoint(
        current_user.unit_id
    )
    if not schedule:
        toast(
            "No active checkpoint currently",
            "warning"
        )
        return redirect(
            url_for(
                "main.dashboard"
            )
        )

    existing = DynamicAttendance.query.filter_by(
        user_id=current_user.id,
        attendance_date=nigeria_now().date(),
        checkpoint_number=schedule.checkpoint_number
    ).first()
    if existing:
        toast(
            f"Checkpoint {schedule.checkpoint_number} already completed",
            "warning"
        )
        return redirect(
            url_for(
                "main.dashboard"
            )
        )
    today = nigeria_now().date()
    todays_break = AttendanceBreak.query.filter_by(
        user_id=current_user.id,
        attendance_date=today
    ).first()
    today = nigeria_now().date()

    completed_checkpoints = DynamicAttendance.query.filter_by(

        user_id=current_user.id,

        attendance_date=today

    ).count()

    todays_break = AttendanceBreak.query.filter_by(

        user_id=current_user.id,

        attendance_date=today

    ).first()

    total_checkpoints = AttendanceSchedule.query.filter_by(

        unit_id=current_user.unit_id,

        is_active=True

    ).count()

    required_checkpoints = total_checkpoints

    if todays_break:

        required_checkpoints -= 1

    compliance = round(

        (
            completed_checkpoints /
            max(1, required_checkpoints)
        ) * 100,

        1

    )
    timeline = []

    schedules = AttendanceSchedule.query.filter_by(

        unit_id=current_user.unit_id,

        is_active=True

    ).order_by(

        AttendanceSchedule.checkpoint_number

    ).all()

    for item in schedules:

        attendance = DynamicAttendance.query.filter_by(

            user_id=current_user.id,

            attendance_date=today,

            checkpoint_number=item.checkpoint_number

        ).first()

        is_break = False

        if todays_break:

            is_break = (

                todays_break.checkpoint_number

                ==

                item.checkpoint_number

            )

        if attendance:

            status = "completed"

        elif is_break:

            status = "break"

        else:

            status = "pending"

        timeline.append({

            "checkpoint": item.checkpoint_number,

            "status": status,

            "open_time": item.open_time,

            "close_time": item.close_time

        })

    remaining_checkpoints = max(

        0,

        required_checkpoints - completed_checkpoints

    )
    return render_template(

        "field_attendance_v2/checkpoint.html",
        schedule=schedule,
        current_time=nigeria_now().strftime("%I:%M %p"),
        completed_checkpoints=completed_checkpoints,
        required_checkpoints=required_checkpoints,
        remaining_checkpoints=remaining_checkpoints,
        compliance=compliance,
        timeline=timeline,
        todays_break=todays_break
    )
#######################################################Break #############################################
@main.route(
    "/field-attendance-v2/break/<int:checkpoint_number>"
)
@login_required
def use_break(checkpoint_number):

    today = nigeria_now().date()

    # Check if user has already used a break today

    existing_break = AttendanceBreak.query.filter_by(

        user_id=current_user.id,

        attendance_date=today

    ).first()
    all_breaks = AttendanceBreak.query.filter_by(
        user_id=current_user.id,
        attendance_date=today
    ).all()
    if existing_break:

        toast(

            f"You already used Checkpoint "
            f"{existing_break.checkpoint_number} "
            f"as your break today",

            "warning"

        )

        return redirect(
            url_for(
                "main.dashboard"
            )
        )

    # Check if user already checked in for this checkpoint

    attendance = DynamicAttendance.query.filter_by(

        user_id=current_user.id,

        attendance_date=today,

        checkpoint_number=checkpoint_number

    ).first()

    if attendance:

        toast(

            "You have already checked in for this checkpoint",

            "warning"

        )

        return redirect(
            url_for(
                "main.dynamic_field_checkpoint"
            )
        )

    # Save break

    attendance_break = AttendanceBreak(

        user_id=current_user.id,

        attendance_date=today,

        checkpoint_number=checkpoint_number

    )

    db.session.add(
        attendance_break
    )

    db.session.commit()

    log_audit(

        "Field Attendance V2",

        "Break",

        f"{current_user.username} used "
        f"Checkpoint {checkpoint_number} "
        f"as break"

    )

    toast(

        f"Checkpoint {checkpoint_number} "
        f"successfully marked as your break",

        "success"

    )

    return redirect(
        url_for(
            "main.dashboard"
        )
    )
@main.route("/field-attendance-v2/checkin", methods=["GET", "POST"])
@login_required
def dynamic_checkin():

    today = nigeria_now().date()

    schedule = get_current_checkpoint(
        current_user.unit_id
    )

    if not schedule:

        toast(
            "No active checkpoint currently",
            "warning"
        )

        return redirect(
            url_for(
                "main.dashboard"
            )
        )

    todays_break = AttendanceBreak.query.filter_by(

        user_id=current_user.id,

        attendance_date=today

    ).first()

    if todays_break and (
        todays_break.checkpoint_number
        ==
        schedule.checkpoint_number
    ):

        toast(

            "This checkpoint has been used as break",

            "warning"

        )

        return redirect(
            url_for(
                "main.dashboard"
            )
        )

    existing = DynamicAttendance.query.filter_by(

        user_id=current_user.id,

        attendance_date=today,

        checkpoint_number=schedule.checkpoint_number

    ).first()

    if existing:

        toast(

            f"Checkpoint "
            f"{schedule.checkpoint_number} "
            f"already completed",

            "warning"

        )

        return redirect(
            url_for(
                "main.dashboard"
            )
        )
    if request.method == "POST":

        photo_data = request.form.get(
            "photo"
        )

        latitude = request.form.get(
            "latitude"
        )

        longitude = request.form.get(
            "longitude"
        )

        remarks = request.form.get(
            "remarks"
        )

        if not latitude or not longitude:

            toast(
                "Location is required",
                "danger"
            )

            return redirect(
                url_for(
                    "main.dynamic_checkin"
                )
            )

        if not photo_data:

            toast(
                "Selfie is required",
                "danger"
            )

            return redirect(
                url_for(
                    "main.dynamic_checkin"
                )
            )

        active_locations = UnitLocation.query.filter_by(

            unit_id=current_user.unit_id,

            status=True

        ).all()

        if not active_locations:

            toast(

                "No attendance location configured",

                "danger"

            )

            return redirect(
                url_for(
                    "main.dashboard"
                )
            )

        allowed = False

        matched_location = None

        matched_distance = None

        for location in active_locations:

            distance = distance_in_meters(

                float(latitude),

                float(longitude),

                location.latitude,

                location.longitude

            )

            if distance <= location.radius:

                allowed = True

                matched_location = location

                matched_distance = distance

                break

        if not allowed:

            log_audit(

                "Field Attendance V2",

                "Failed Check-In",

                f"{current_user.username} attempted "
                f"Checkpoint "
                f"{schedule.checkpoint_number} "
                f"outside approved location"

            )

            toast(
                    f"You are outside the approved attendance location"
                    f"{round(distance,2)}m "
                    f"away from "
                    f'{"SUBMITTED LATITUDE:", latitude}'
                    f'{"SUBMITTED LONGITUDE:", longitude}'
                    f"{location.name}. "
                    f"Allowed radius is "
                    f"{location.radius}m",
                    "danger"

                )
            return redirect(
                url_for(
                    "main.dynamic_checkin"
                )
            )

        upload_folder = os.path.join(

            "app",

            "static",

            "uploads",

            "dynamic_attendance"

        )

        os.makedirs(

            upload_folder,

            exist_ok=True

        )

        filename = (

            f"dyn_"

            f"{current_user.id}_"

            f"{today}_"

            f"{schedule.checkpoint_number}.jpg"

        )

        filepath = os.path.join(

            upload_folder,

            filename

        )

        header, encoded = photo_data.split(
            ",",
            1
        )

        with open(
            filepath,
            "wb"
        ) as f:

            f.write(
                base64.b64decode(
                    encoded
                )
            )

        attendance = DynamicAttendance(

            user_id=current_user.id,

            unit_id=current_user.unit_id,

            location_id=matched_location.id,

            checkpoint_number=schedule.checkpoint_number,

            attendance_date=today,

            attendance_time=nigeria_now(),

            latitude=float(latitude),

            longitude=float(longitude),

            photo=filename,

            distance=matched_distance,

            remarks=remarks,

            status="Present"

        )

        db.session.add(
            attendance
        )

        db.session.commit()

        log_audit(

            "Field Attendance V2",

            "Check-In",

            f"{current_user.username} completed "
            f"Checkpoint "
            f"{schedule.checkpoint_number} "
            f" at "
            f"{matched_location.name} "
            f"({round(matched_distance, 2)}m)"

        )

        toast(

            f"Checkpoint "
            f"{schedule.checkpoint_number} "
            f"recorded successfully",

            "success"

        )

        return redirect(
            url_for(
                "main.dashboard"
            )
        )
    return render_template(

        "field_attendance_v2/checkin.html",

        schedule=schedule,

        current_time=nigeria_now().strftime(
            "%I:%M %p"
        ),

        todays_break=todays_break
    )
@main.route(
    "/field-attendance-v2/history"
)
@login_required
def attendance_history_v2():

    start_date = request.args.get(
        "start_date"
    )

    end_date = request.args.get(
        "end_date"
    )
    records_query = DynamicAttendance.query.filter_by(

        user_id=current_user.id

    )

    breaks_query = AttendanceBreak.query.filter_by(

        user_id=current_user.id

    )

    if start_date:

        start_dt = datetime.strptime(
            start_date,
            "%Y-%m-%d"
        ).date()

        records_query = records_query.filter(
            DynamicAttendance.attendance_date >= start_dt
        )

        breaks_query = breaks_query.filter(
            AttendanceBreak.attendance_date >= start_dt
        )

    if end_date:

        end_dt = datetime.strptime(
            end_date,
            "%Y-%m-%d"
        ).date()

        records_query = records_query.filter(
            DynamicAttendance.attendance_date <= end_dt
        )

        breaks_query = breaks_query.filter(
            AttendanceBreak.attendance_date <= end_dt
        )

    records = records_query.order_by(

        DynamicAttendance.attendance_date.desc(),

        DynamicAttendance.attendance_time.desc()

    ).all()

    breaks = breaks_query.order_by(

        AttendanceBreak.attendance_date.desc()

    ).all()

    total_attendance = len(
        records
    )

    total_breaks = len(
        breaks
    )

    log_audit(

        "Field Attendance V2",

        "History",

        f"{current_user.username} viewed attendance history"

    )

    return render_template(

        "field_attendance_v2/history.html",

        records=records,

        breaks=breaks,

        total_attendance=total_attendance,

        total_breaks=total_breaks,

        start_date=start_date,

        end_date=end_date

    )
@main.route(
    "/field-attendance-v2/compliance"
)
@login_required
def attendance_compliance_v2():

    selected_date = request.args.get(
        "report_date"
    )

    user_id = request.args.get(
        "user_id"
    )

    unit_id = request.args.get(
        "unit_id"
    )

    if selected_date:

        report_date = datetime.strptime(

            selected_date,

            "%Y-%m-%d"

        ).date()

    else:

        report_date = nigeria_now().date()

    # ==================================
    # STAFF FILTER
    # ==================================

    staff_query = User.query.filter_by(

        is_active_user=True

    )

    if unit_id:

        staff_query = staff_query.filter(

            User.unit_id == unit_id

        )

    if user_id:

        staff_query = staff_query.filter(

            User.id == user_id

        )

    staff = staff_query.order_by(

        User.full_name

    ).all()

    # ==================================
    # BUILD REPORT
    # ==================================

    compliance_data = []

    for user in staff:

        punches = DynamicAttendance.query.filter_by(

            user_id=user.id,

            attendance_date=report_date

        ).all()

        breaks = AttendanceBreak.query.filter_by(

            user_id=user.id,

            attendance_date=report_date

        ).all()

        completed = len(punches)

        config = AttendanceScheduleConfig.query.filter_by(

            unit_id=user.unit_id).first()

        if config:

            expected = config.total_checkpoints

        else:

            expected = 0

        if breaks:

            expected = max(

                0,

                expected - 1

            )

        compliance = round(

            (
                completed /
                max(1, expected)
            ) * 100,

            1

        )
        if compliance >= 90:

            status = "Excellent"

        elif compliance >= 75:

            status = "Good"

        elif compliance >= 50:

            status = "Fair"

        else:

            status = "Poor"

        compliance_data.append({

            "staff_name": user.full_name,

            "project": (
                user.unit.name
                if user.unit
                else ""
            ),

            "completed": completed,

            "expected": expected,

            "break_used": (
                "Yes"
                if breaks
                else "No"
            ),

            "compliance": compliance,

            "status": status

        })

    # ==================================
    # DROPDOWNS
    # ==================================

    units = Unit.query.filter_by(

        status=True

    ).order_by(

        Unit.name

    ).all()

    users = User.query.filter_by(

        is_active_user=True

    ).order_by(

        User.full_name

    ).all()

    log_audit(

        "Field Attendance V2",

        "Compliance Report",

        f"{current_user.username} viewed compliance report"

    )

    return render_template(

        "field_attendance_v2/compliance.html",

        compliance_data=compliance_data,

        report_date=report_date,

        users=users,

        units=units,

        user_id=user_id,

        unit_id=unit_id

    )


@main.route(
    "/field-attendance-v2/dashboard"
)
@login_required
@admin_required
def attendance_dashboard_v2():

    # =====================================
    # FILTERS
    # =====================================

    selected_date = request.args.get(
        "report_date"
    )

    unit_id = request.args.get(
        "unit_id"
    )

    user_id = request.args.get(
        "user_id"
    )

    status_filter = request.args.get(
        "status"
    )

    if selected_date:

        today = datetime.strptime(

            selected_date,

            "%Y-%m-%d"

        ).date()

    else:

        today = nigeria_now().date()

    # =====================================
    # KPI COUNTERS
    # =====================================

    total_staff = User.query.filter_by(
        is_active_user=True
    ).count()

    excellent_count = 0
    good_count = 0
    fair_count = 0
    poor_count = 0

    # =====================================
    # PROJECT SUMMARY
    # =====================================

    units = Unit.query.filter_by(
        status=True
    ).order_by(
        Unit.name
    ).all()

    project_summary = []

    for unit in units:

        staff_members = User.query.filter_by(

            unit_id=unit.id,

            is_active_user=True

        ).all()

        staff_count = len(
            staff_members
        )

        total_compliance = 0

        for staff in staff_members:

            punches = DynamicAttendance.query.filter_by(

                user_id=staff.id,

                attendance_date=today

            ).count()

            breaks = AttendanceBreak.query.filter_by(

                user_id=staff.id,

                attendance_date=today

            ).count()

            config = AttendanceScheduleConfig.query.filter_by(

                unit_id=staff.unit_id

            ).first()

            expected = (

                config.total_checkpoints

                if config

                else 0

            )

            if breaks:

                expected = max(
                    0,
                    expected - 1
                )

            compliance = round(

                (
                    punches /
                    max(1, expected)
                ) * 100,

                1

            )

            total_compliance += compliance

            if compliance >= 90:

                excellent_count += 1

            elif compliance >= 75:

                good_count += 1

            elif compliance >= 50:

                fair_count += 1

            else:

                poor_count += 1

        avg_compliance = round(

            total_compliance /
            max(1, staff_count),

            1

        )

        project_summary.append({

            "project": unit.name,

            "staff_count": staff_count,

            "avg_compliance": avg_compliance

        })

    # =====================================
    # PROJECT PAGINATION
    # =====================================

    project_page = request.args.get(

        "project_page",

        1,

        type=int

    )

    projects_per_page = 5

    project_total = len(
        project_summary
    )

    project_total_pages = math.ceil(

        project_total /

        projects_per_page

    )

    start = (

        (project_page - 1)

        * projects_per_page

    )

    end = start + projects_per_page

    project_summary = project_summary[
        start:end
    ]

    # =====================================
    # STAFF QUERY
    # =====================================

    staff_query = User.query.filter_by(

        is_active_user=True

    )

    if unit_id:

        staff_query = staff_query.filter(

            User.unit_id == unit_id

        )

    if user_id:

        staff_query = staff_query.filter(

            User.id == user_id

        )

    staff_query = staff_query.order_by(

        User.full_name

    )

    page = request.args.get(

        "page",

        1,

        type=int

    )

    staff_pagination = staff_query.paginate(

        page=page,

        per_page=10,

        error_out=False

    )

    # =====================================
    # STAFF COMPLIANCE
    # =====================================

    staff_compliance = []

    for staff in staff_pagination.items:

        punches = DynamicAttendance.query.filter_by(

            user_id=staff.id,

            attendance_date=today

        ).count()

        breaks = AttendanceBreak.query.filter_by(

            user_id=staff.id,

            attendance_date=today

        ).count()

        config = AttendanceScheduleConfig.query.filter_by(

            unit_id=staff.unit_id

        ).first()

        expected = (

            config.total_checkpoints

            if config

            else 0

        )

        if breaks:

            expected = max(
                0,
                expected - 1
            )

        compliance = round(

            (
                punches /
                max(1, expected)
            ) * 100,

            1

        )

        if compliance >= 90:

            status = "Excellent"

        elif compliance >= 75:

            status = "Good"

        elif compliance >= 50:

            status = "Fair"

        else:

            status = "Poor"

        if status_filter:

            if status != status_filter:

                continue

        staff_compliance.append({

            "staff_name": staff.full_name,

            "project": (
                staff.unit.name
                if staff.unit
                else ""
            ),

            "completed": punches,

            "expected": expected,

            "compliance": compliance,

            "status": status

        })

    # =====================================
    # DROPDOWNS
    # =====================================

    users = User.query.filter_by(

        is_active_user=True

    ).order_by(

        User.full_name

    ).all()
    total_compliance = (

        excellent_count * 100 +

        good_count * 75 +

        fair_count * 50 +

        poor_count * 25

    )

    average_compliance = round(

        total_compliance /

        max(
            1,
            total_staff
        ),

        1

    )

    # =====================================
    # AUDIT LOG
    # =====================================

    log_audit(

        "Field Attendance V2",

        "Dashboard",

        f"{current_user.username} viewed dashboard"

    )

    # =====================================
    # RENDER
    # =====================================

    return render_template(

        "field_attendance_v2/dashboard.html",

        total_staff=total_staff,

        excellent_count=excellent_count,

        good_count=good_count,

        fair_count=fair_count,

        poor_count=poor_count,

        today=today,
        average_compliance=average_compliance,
        report_date=today,

        project_summary=project_summary,

        project_page=project_page,

        project_total_pages=project_total_pages,

        staff_compliance=staff_compliance,

        staff_pagination=staff_pagination,

        units=units,

        users=users,

        unit_id=unit_id,

        user_id=user_id,

        status_filter=status_filter

    )

@main.route(
    "/field-attendance-v2/dashboard/excel"
)
@login_required
@admin_required
def dashboard_export_excel():

    selected_date = request.args.get(
        "report_date"
    )

    unit_id = request.args.get(
        "unit_id"
    )

    user_id = request.args.get(
        "user_id"
    )

    if selected_date:

        report_date = datetime.strptime(
            selected_date,
            "%Y-%m-%d"
        ).date()

    else:

        report_date = nigeria_now().date()

    wb = Workbook()

    ws = wb.active

    ws.title = "Dashboard Report"

    headers = [

        "Staff",

        "Project",

        "Completed",

        "Expected",

        "Compliance %",

        "Status"

    ]

    for col_num, header in enumerate(
        headers,
        start=1
    ):

        ws.cell(
            row=1,
            column=col_num,
            value=header
        )

    row_num = 2

    staff_query = User.query.filter_by(
        is_active_user=True
    )

    if unit_id:

        staff_query = staff_query.filter(
            User.unit_id == unit_id
        )

    if user_id:

        staff_query = staff_query.filter(
            User.id == user_id
        )

    for staff in staff_query.all():

        punches = DynamicAttendance.query.filter_by(

            user_id=staff.id,

            attendance_date=report_date

        ).count()

        breaks = AttendanceBreak.query.filter_by(

            user_id=staff.id,

            attendance_date=report_date

        ).count()

        config = AttendanceScheduleConfig.query.filter_by(

            unit_id=staff.unit_id

        ).first()

        expected = (

            config.total_checkpoints

            if config

            else 0

        )

        if breaks:

            expected = max(
                0,
                expected - 1
            )

        compliance = round(

            (
                punches /
                max(1, expected)
            ) * 100,

            1

        )

        if compliance >= 90:

            status = "Excellent"

        elif compliance >= 75:

            status = "Good"

        elif compliance >= 50:

            status = "Fair"

        else:

            status = "Poor"

        ws.append([

            staff.full_name,

            staff.unit.name
            if staff.unit
            else "",

            punches,

            expected,

            compliance,

            status

        ])

        row_num += 1

    output = BytesIO()

    wb.save(
        output
    )

    output.seek(0)

    return send_file(

        output,

        download_name=(
            f"dashboard_report_"
            f"{report_date}.xlsx"
        ),

        as_attachment=True,

        mimetype=(
            "application/"
            "vnd.openxmlformats-"
            "officedocument."
            "spreadsheetml.sheet"
        )

    )
@main.route(
    "/field-attendance-v2/dashboard/pdf"
)
@login_required
@admin_required
def dashboard_export_pdf():

    selected_date = request.args.get(
        "report_date"
    )

    unit_id = request.args.get(
        "unit_id"
    )

    user_id = request.args.get(
        "user_id"
    )

    status_filter = request.args.get(
        "status"
    )

    if selected_date:

        report_date = datetime.strptime(
            selected_date,
            "%Y-%m-%d"
        ).date()

    else:

        report_date = nigeria_now().date()

    # ==========================
    # STAFF FILTERS
    # ==========================

    staff_query = User.query.filter_by(
        is_active_user=True
    )

    if unit_id:

        staff_query = staff_query.filter(
            User.unit_id == unit_id
        )

    if user_id:

        staff_query = staff_query.filter(
            User.id == user_id
        )

    staff_query = staff_query.order_by(
        User.full_name
    )

    # ==========================
    # PDF
    # ==========================

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(

        Paragraph(

            f"WorkForce Dashboard Report - {report_date}",

            styles["Heading1"]

        )

    )

    elements.append(
        Spacer(1, 12)
    )

    data = [[

        "Staff",

        "Project",

        "Completed",

        "Expected",

        "Compliance %",

        "Status"

    ]]

    for staff in staff_query.all():

        punches = DynamicAttendance.query.filter_by(

            user_id=staff.id,

            attendance_date=report_date

        ).count()

        breaks = AttendanceBreak.query.filter_by(

            user_id=staff.id,

            attendance_date=report_date

        ).count()

        config = AttendanceScheduleConfig.query.filter_by(

            unit_id=staff.unit_id

        ).first()

        expected = (

            config.total_checkpoints

            if config

            else 0

        )

        if breaks:

            expected = max(
                0,
                expected - 1
            )

        compliance = round(

            (
                punches /
                max(1, expected)
            ) * 100,

            1

        )

        if compliance >= 90:

            status = "Excellent"

        elif compliance >= 75:

            status = "Good"

        elif compliance >= 50:

            status = "Fair"

        else:

            status = "Poor"

        # ==========================
        # STATUS FILTER
        # ==========================

        if status_filter:

            if status != status_filter:

                continue

        data.append([

            staff.full_name,

            staff.unit.name
            if staff.unit
            else "",

            str(punches),

            str(expected),

            f"{compliance}%",

            status

        ])

    table = Table(data)

    table.setStyle(

        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightblue
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.black
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                1,
                colors.black
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            )

        ])

    )

    elements.append(
        table
    )

    doc.build(
        elements
    )

    buffer.seek(0)

    log_audit(

        "Field Attendance V2",

        "PDF Export",

        f"{current_user.username} exported dashboard PDF"

    )

    return send_file(

        buffer,

        as_attachment=True,

        download_name=(

            f"dashboard_report_"

            f"{report_date}.pdf"

        ),

        mimetype="application/pdf"

    )

##############################BA Attaendance History########################################
@main.route(
    "/field-attendance-v2/all_history"
)
@login_required
@admin_required
def all_history():

    start_date = request.args.get(
        "start_date"
    )

    end_date = request.args.get(
        "end_date"
    )

    unit_id = request.args.get(
        "unit_id"
    )

    user_id = request.args.get(
        "user_id"
    )

    page = request.args.get(

        "page",

        1,

        type=int

    )

    records_query = db.session.query(

        DynamicAttendance

    ).join(

        User,

        User.id == DynamicAttendance.user_id

    )

    breaks_query = db.session.query(

        AttendanceBreak

    ).join(

        User,

        User.id == AttendanceBreak.user_id

    )

    if unit_id:

        records_query = records_query.filter(

            User.unit_id == unit_id

        )

        breaks_query = breaks_query.filter(

            User.unit_id == unit_id

        )

    if user_id:

        records_query = records_query.filter(

            DynamicAttendance.user_id == user_id

        )

        breaks_query = breaks_query.filter(

            AttendanceBreak.user_id == user_id

        )

    if start_date:

        start_dt = datetime.strptime(

            start_date,

            "%Y-%m-%d"

        ).date()

        records_query = records_query.filter(

            DynamicAttendance.attendance_date >= start_dt

        )

        breaks_query = breaks_query.filter(

            AttendanceBreak.attendance_date >= start_dt

        )

    if end_date:

        end_dt = datetime.strptime(

            end_date,

            "%Y-%m-%d"

        ).date()

        records_query = records_query.filter(

            DynamicAttendance.attendance_date <= end_dt

        )

        breaks_query = breaks_query.filter(

            AttendanceBreak.attendance_date <= end_dt

        )

    records = records_query.order_by(

        DynamicAttendance.attendance_date.desc(),

        DynamicAttendance.attendance_time.desc()

    ).paginate(

        page=page,

        per_page=20,

        error_out=False

    )

    breaks = breaks_query.order_by(

        AttendanceBreak.attendance_date.desc()

    ).all()

    total_attendance = records.total

    total_breaks = len(

        breaks

    )

    projects = Unit.query.filter_by(

        status=True

    ).order_by(

        Unit.name

    ).all()

    users_query = User.query.filter_by(

        is_active_user=True

    )

    if unit_id:

        users_query = users_query.filter(

            User.unit_id == unit_id

        )

    users = users_query.order_by(

        User.full_name

    ).all()

    log_audit(

        "Field Attendance V2",

        "History",

        f"{current_user.username} viewed attendance history"

    )

    return render_template(

        "field_attendance_v2/ba_history.html",

        records=records,

        breaks=breaks,

        total_attendance=total_attendance,

        total_breaks=total_breaks,

        start_date=start_date,

        end_date=end_date,

        projects=projects,

        users=users,

        unit_id=unit_id,

        user_id=user_id

    )

@main.route(

    "/api/project_users/<int:unit_id>"

)

@login_required
def project_users(unit_id):

    users = User.query.filter_by(

        unit_id=unit_id,

        is_active_user=True

    ).order_by(

        User.full_name

    ).all()

    return jsonify([

        {

            "id": u.id,

            "name": u.full_name

        }

        for u in users

    ])

####################################Export History to Excel########################################
def build_history_queryset():

    start_date = request.args.get(
        "start_date"
    )

    end_date = request.args.get(
        "end_date"
    )

    unit_id = request.args.get(
        "unit_id"
    )

    user_id = request.args.get(
        "user_id"
    )

    query = DynamicAttendance.query

    if unit_id:

        query = query.join(

            User

        ).filter(

            User.unit_id == unit_id

        )

    if user_id:

        query = query.filter(

            DynamicAttendance.user_id == user_id

        )

    if start_date:

        start_dt = datetime.strptime(

            start_date,

            "%Y-%m-%d"

        ).date()

        query = query.filter(

            DynamicAttendance.attendance_date >= start_dt

        )

    if end_date:

        end_dt = datetime.strptime(

            end_date,

            "%Y-%m-%d"

        ).date()

        query = query.filter(

            DynamicAttendance.attendance_date <= end_dt

        )

    return query.order_by(

        DynamicAttendance.attendance_date.desc(),

        DynamicAttendance.attendance_time.desc()

    )

@main.route(
    "/field-attendance-v2/history/excel"
)
@login_required
@admin_required
def export_history_excel():

    records = build_history_queryset().all()

    wb = Workbook()

    ws = wb.active

    ws.title = "Attendance"

    headers = [

        "Staff",

        "Project",

        "Date",

        "Checkpoint",

        "Time",

        "Distance",

        "Remarks"

    ]

    ws.append(

        headers

    )

    for item in records:

        ws.append([

            item.users.full_name,

            item.users.unit.name
            if item.users.unit
            else "",

            str(

                item.attendance_date

            ),

            item.checkpoint_number,

            item.attendance_time.strftime(

                '%I:%M %p'

            ),

            item.distance,

            item.remarks

        ])

    output = BytesIO()

    wb.save(

        output

    )

    output.seek(

        0

    )

    return send_file(

        output,

        as_attachment=True,

        download_name=

        "attendance.xlsx",

        mimetype=

        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

@main.route(
"/field-attendance-v2/history/pdf"
)
@login_required
@admin_required
def export_history_pdf():

    records = build_history_queryset().all()

    buffer = BytesIO()

    doc = SimpleDocTemplate(

        buffer,

        pagesize=landscape(A4)

    )

    elements = []

    data = [[

        "Staff",

        "Project",

        "Date",

        "CP",

        "Time",

        "Distance"

    ]]

    for item in records:

        data.append([

            item.users.full_name,

            item.users.unit.name
            if item.users.unit
            else "",

            str(

                item.attendance_date

            ),

            item.checkpoint_number,

            item.attendance_time.strftime(

                '%I:%M %p'

            ),

            str(

                item.distance

            )

        ])

    table = Table(

        data

    )

    table.setStyle(

        TableStyle([

            (

                'BACKGROUND',

                (0,0),

                (-1,0),

                colors.lightblue

            ),

            (

                'GRID',

                (0,0),

                (-1,-1),

                1,

                colors.black

            ),

            (

                'FONTSIZE',

                (0,0),

                (-1,-1),

                8

            )

        ])

    )

    elements.append(

        table

    )

    doc.build(

        elements

    )

    buffer.seek(

        0

    )

    return send_file(

        buffer,

        as_attachment=True,

        download_name=

        "attendance.pdf",

        mimetype=

        "application/pdf"

    )

@main.route("/users/import", methods=["GET", "POST"])
@login_required
@admin_required
def import_users():

    if request.method == "POST":

        if "excel_file" not in request.files:

            toast("Please select an Excel file.", "warning")
            return redirect(url_for("main.import_users"))

        file = request.files["excel_file"]

        if file.filename == "":

            toast("Please select an Excel file.", "warning")
            return redirect(url_for("main.import_users"))

        try:

            workbook = load_workbook(file)
            sheet = workbook["User Import"]

        #except Exception:

            #toast("Invalid Excel template.", "danger")
            #return redirect(url_for("main.import_users"))
        except Exception as e:

            toast(str(e), "danger")

            print(e)

            return redirect(url_for("main.import_users"))

        imported = 0
        skipped = 0
        errors = []

        for row_number, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True),
                start=2):

            # Skip completely empty rows
            if not any(row):
                continue

            full_name = str(row[0]).strip() if row[0] else ""
            username = str(row[1]).strip().lower() if row[1] else ""
            email = str(row[2]).strip().lower() if row[2] else ""
            phone = str(row[3]).strip() if row[3] else ""
            project_name = str(row[4]).strip() if row[4] else ""
            department_name = str(row[5]).strip() if row[5] else ""
            role = str(row[6]).strip().lower() if row[6] else "staff"

            # ==========================
            # Validation
            # ==========================

            if full_name == "":
                skipped += 1
                errors.append(f"Row {row_number}: Full Name is missing.")
                continue

            if username == "":
                skipped += 1
                errors.append(f"Row {row_number}: Username is missing.")
                continue

            if User.query.filter_by(username=username).first():

                skipped += 1
                errors.append(
                    f"Row {row_number}: Username '{username}' already exists."
                )
                continue

            unit = Unit.query.filter_by(
                name=project_name
            ).first()

            if not unit:

                skipped += 1
                errors.append(
                    f"Row {row_number}: Project '{project_name}' not found."
                )
                continue

            department = Department.query.filter_by(
                name=department_name
            ).first()

            if not department:

                skipped += 1
                errors.append(
                    f"Row {row_number}: Department '{department_name}' not found."
                )
                continue

            staff_id = generate_staff_id()

            user = User(

                staff_id=staff_id,

                full_name=full_name,

                username=username,

                email=email,

                phone=phone,

                unit_id=unit.id,

                department_id=department.id,

                role=role

            )

            user.set_password("12345")

            db.session.add(user)

            imported += 1

        try:

            db.session.commit()

        except Exception as e:

            db.session.rollback()

            toast(
                f"Database Error: {e}",
                "danger"
            )

            return redirect(
                url_for("main.import_users")
            )

        log_audit(
            "Users",
            "Bulk Import",
            f"Imported {imported} users"
        )

        return render_template(
            "users/import.html",
            imported=imported,
            skipped=skipped,
            errors=errors
        )

    return render_template(
        "users/import.html"
    )


@main.route("/users/download-template")
@login_required
@admin_required
def download_user_template():

    wb = Workbook()

    # ===========================
    # User Import Sheet
    # ===========================
    ws = wb.active
    ws.title = "User Import"

    headers = [
        "Full Name",
        "Username",
        "Email",
        "Phone",
        "Project",
        "Department",
        "Role"
    ]

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(row=1, column=col)

        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Freeze header
    ws.freeze_panes = "A2"

    # Filter
    ws.auto_filter.ref = "A1:G500"

    # Sample Row
    ws.append([
        "John Doe",
        "jdoe",
        "john@example.com",
        "08031234567",
        "",
        "",
        ""
    ])

    # Column Widths
    widths = {
        "A": 30,
        "B": 20,
        "C": 35,
        "D": 18,
        "E": 25,
        "F": 25,
        "G": 20
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    # ===========================
    # Projects Sheet
    # ===========================
    project_sheet = wb.create_sheet("Projects")

    project_sheet.append(["Available Projects"])

    projects = Unit.query.filter_by(status=True)\
                         .order_by(Unit.name)\
                         .all()

    for project in projects:
        project_sheet.append([project.name])

    # ===========================
    # Departments Sheet
    # ===========================
    dept_sheet = wb.create_sheet("Departments")

    dept_sheet.append(["Available Departments"])

    departments = Department.query.filter_by(status=True)\
                                  .order_by(Department.name)\
                                  .all()

    for dept in departments:
        dept_sheet.append([dept.name])

    # ===========================
    # Roles Sheet
    # ===========================
    role_sheet = wb.create_sheet("Roles")

    role_sheet.append(["Available Roles"])

    roles = [
        "superadmin",
        "admin",
        "supervisor",
        "staff"
    ]

    for role in roles:
        role_sheet.append([role])

    # ===========================
    # Excel Drop-down Lists
    # ===========================
    project_validation = DataValidation(
        type="list",
        formula1="=Projects!$A$2:$A$500",
        allow_blank=False
    )

    department_validation = DataValidation(
        type="list",
        formula1="=Departments!$A$2:$A$500",
        allow_blank=False
    )

    role_validation = DataValidation(
        type="list",
        formula1="=Roles!$A$2:$A$5",
        allow_blank=False
    )

    ws.add_data_validation(project_validation)
    ws.add_data_validation(department_validation)
    ws.add_data_validation(role_validation)

    project_validation.add("E2:E500")
    department_validation.add("F2:F500")
    role_validation.add("G2:G500")

    # ===========================
    # Download Workbook
    # ===========================
    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="WorkForce_User_Import_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )